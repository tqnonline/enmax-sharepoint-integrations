"""Extract master/reference data from Master data.xlsx to YAML seed files.

Reads the ColumnValues sheet, applies five transformations from PRD section 22,
and emits per-table YAML to solution/seed/reference/.

Five transformations (applied in order per cell):
  1. Code / display-name split on first " - "
  2. < → under, > → over for System/Shepard columns
  3. Encoding cleanup: ¿ → – (en dash)
  4. Deduplication on code (first occurrence wins)
  5. XXX / XX sentinel → display_name = "Unspecified"

Usage:
    python solution/scripts/extract_master_data.py [--workbook <path>]
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import yaml

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: uv pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent


def _main_repo_root() -> Path:
    """Walk up from REPO_ROOT to find the dir that owns the .worktrees/ folder.

    When running inside a worktree (e.g. .worktrees/feat-002/), REPO_ROOT points
    at the worktree directory, not the main repository root.  The main repo root
    is the first ancestor that contains a .worktrees/ directory.
    """
    p = REPO_ROOT
    while p != p.parent:
        if (p / ".worktrees").is_dir():
            return p
        p = p.parent
    return REPO_ROOT


_MAIN_REPO = _main_repo_root()
DEFAULT_WORKBOOK = (
    _MAIN_REPO
    / ".worktrees"
    / "specs"
    / "docs"
    / "superpowers"
    / "specs"
    / "_assets"
    / "master-data"
    / "Master data.xlsx"
)
SEED_REF_DIR = REPO_ROOT / "solution" / "seed" / "reference"

# ColumnValues sheet layout
_HEADERS_ROW = 3   # 1-indexed
_DATA_START = 4    # 1-indexed

# Columns where < → under and > → over substitution applies
_ANGLE_BRACKET_COLS = frozenset({"System", "Shepard"})

# Sentinel codes that always survive dedup with display_name = "Unspecified"
_SENTINELS = frozenset({"XXX", "XX"})

# Excel encoding artefact: ¿ (U+00BF) was mis-decoded from – (en dash U+2013)
_ENCODING_FIX = {"¿": "–"}


# ---------------------------------------------------------------------------
# Cell-level transformations
# ---------------------------------------------------------------------------

def _fix_encoding(text: str) -> str:
    for bad, good in _ENCODING_FIX.items():
        text = text.replace(bad, good)
    return text


# Ensure a space after "under"/"over" when immediately followed by a non-space
# e.g. ColumnValueMapping may produce "under1000V" from "<1000V"
_UNDER_OVER_SPACE_RE = re.compile(r"\b(under|over)(\S)")


def _angle_sub(text: str) -> str:
    text = text.replace("<", "under ").replace(">", "over ")
    # Collapse any accidental double spaces introduced by the substitution
    return re.sub(r"  +", " ", text).strip()


def _fix_under_over_spacing(text: str) -> str:
    """Add missing space in already-substituted values like 'under1000V'."""
    return _UNDER_OVER_SPACE_RE.sub(r"\1 \2", text)


def _parse_cell(
    raw: Any, col_header: str, cvm: dict[str, str]
) -> tuple[str, str] | None:
    """Return (code, display_name) or None for blank/null cells."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None

    # Transformation 3: encoding cleanup
    text = _fix_encoding(text)

    # ColumnValueMapping explicit substitution (precedes general angle-bracket rule)
    if text in cvm:
        text = cvm[text]
        # CVM values may already have under/over substituted but missing a space
        text = _fix_under_over_spacing(text)

    # Transformation 2: angle-bracket columns
    if col_header in _ANGLE_BRACKET_COLS:
        text = _angle_sub(text)
    # Always fix missing space after "under"/"over" for angle-bracket columns —
    # Excel cells may already carry pre-substituted values like "under1000V"
    if col_header in _ANGLE_BRACKET_COLS or "under" in text or "over" in text:
        text = _fix_under_over_spacing(text)

    # Transformation 1: split on first " - "
    if " - " in text:
        code, _, display = text.partition(" - ")
        code = code.strip()
        display = display.strip()
    else:
        code = text
        display = text

    # Transformation 5: XXX / XX sentinels
    if code in _SENTINELS:
        display = "Unspecified"

    return code, display


# ---------------------------------------------------------------------------
# ColumnValueMapping loader
# ---------------------------------------------------------------------------

def _load_cvm(wb) -> dict[str, str]:
    if "ColumnValueMapping" not in wb.sheetnames:
        return {}
    ws = wb["ColumnValueMapping"]
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        old = str(row[0]).strip()
        new = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if old and new:
            result[old] = new
    return result


# ---------------------------------------------------------------------------
# YAML writer
# ---------------------------------------------------------------------------

def _write_yaml(
    path: Path,
    table: str,
    natural_keys: list[str],
    lookups: dict | None,
    rows: list[dict],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    data: dict = {"table": table, "natural_key_columns": natural_keys}
    if lookups:
        data["lookups"] = lookups
    data["rows"] = rows
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        yaml.dump(data, f, allow_unicode=True, sort_keys=False, default_flow_style=False)
    print(f"  {path.name}: {len(rows)} rows")


# ---------------------------------------------------------------------------
# Main single-pass extraction
# ---------------------------------------------------------------------------

def _extract(
    cv_ws,
    header_to_idx: dict[str, int],
    cvm: dict[str, str],
) -> dict[str, list[dict]]:
    """One pass over ColumnValues data rows; returns per-table row lists."""

    # Each column in the Excel is an independent value list (not relational).
    # Parent relationships (asset→business, unit→asset) cannot be inferred here.
    # Junction tables (system_scope) ship empty; populate manually if needed.
    buckets: dict[str, list[dict]] = {
        "business": [], "asset": [], "unit": [],
        "domain": [], "system": [], "kind": [],
        "vendor": [], "record_type": [], "record_phase": [],
    }
    seen: dict[str, set] = {k: set() for k in buckets}
    sort_ctr: dict[str, int] = {k: 10 for k in buckets}

    # Actual Excel column headers (may differ from logical names)
    _COL_ALIAS = {
        "Asset (Facility)": "asset",
        "Domain (Group)": "domain",
    }
    _COL_DIRECT = {
        "Business": "business",
        "Unit": "unit",
        "System": "system",
        "Kind": "kind",
        "Vendor Name": "vendor",
        "Record Type": "record_type",
        "Record Phase": "record_phase",
    }

    def _add(key: str, code: str, display: str) -> bool:
        if code in seen[key]:
            return False
        seen[key].add(code)
        buckets[key].append({
            "code": code,
            "display_name": display,
            "status": 1,  # 1 = Active (enmax_acdn_recordstatus); integer avoids cross-option-set label ambiguity
            "sort_order": sort_ctr[key],
        })
        sort_ctr[key] += 10
        return True

    for row in cv_ws.iter_rows(min_row=_DATA_START, values_only=True):
        # Resolve all headers in this row
        for hdr, idx in header_to_idx.items():
            if idx >= len(row):
                continue
            key = _COL_DIRECT.get(hdr) or _COL_ALIAS.get(hdr)
            if not key:
                continue
            p = _parse_cell(row[idx], hdr, cvm)
            if p:
                _add(key, p[0], p[1])

    return buckets


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract master data from Excel workbook to YAML seed files"
    )
    parser.add_argument(
        "--workbook",
        default=str(DEFAULT_WORKBOOK),
        help="Path to Master data.xlsx",
    )
    args = parser.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f"ERROR: Workbook not found: {wb_path}", file=sys.stderr)
        return 1

    print(f"Loading {wb_path.name}...")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    cvm = _load_cvm(wb)
    print(f"  ColumnValueMapping: {len(cvm)} substitutions")

    cv_ws = wb["ColumnValues"]

    # Discover headers (row 3, 1-indexed)
    headers: list[str] = []
    for i, row in enumerate(cv_ws.iter_rows(min_row=1, max_row=_HEADERS_ROW, values_only=True)):
        if i == _HEADERS_ROW - 1:
            headers = [str(c).strip() if c is not None else "" for c in row]
    header_to_idx = {h: i for i, h in enumerate(headers) if h}
    print(f"  Columns found: {[h for h in header_to_idx if h]}")

    extracted = _extract(cv_ws, header_to_idx, cvm)
    wb.close()

    print("\nExtracted:")
    for key, rows in extracted.items():
        print(f"  {key}: {len(rows)}")

    print(f"\nWriting to {SEED_REF_DIR} ...")

    # Reference tables (code-only natural keys; no FK lookups needed for simple tables)
    _simple: list[tuple[str, str]] = [
        ("business.yaml",    "enmax_autocadbusiness"),
        ("asset.yaml",       "enmax_autocadasset"),
        ("unit.yaml",        "enmax_autocadunit"),
        ("domain.yaml",      "enmax_autocaddomain"),
        ("system.yaml",      "enmax_autocadsystem"),
        ("kind.yaml",        "enmax_autocadkind"),
        ("record_type.yaml", "enmax_autocadrecordtype"),
        ("record_phase.yaml","enmax_autocadrecordphase"),
        ("vendor.yaml",      "enmax_autocadvendor"),
    ]
    for filename, table in _simple:
        key = filename.replace(".yaml", "")
        _write_yaml(SEED_REF_DIR / filename, table, ["code"], None, extracted[key])

    # Junction table ships empty; populate manually if system scoping rules are needed.
    for filename, table, nk_cols in [
        ("system_scope.yaml",   "enmax_autocadsystemscope",   ["system_code", "scope_type", "scope_value"]),
    ]:
        _write_yaml(SEED_REF_DIR / filename, table, nk_cols, None, [])

    print("\nExtraction complete.")
    print("NOTE: system_scope.yaml is an empty stub — populate manually if needed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
