"""Parse legacy GF / EEC coded-PDF Excel extracts for Dataverse import.

Filename, dedupe, Active-only (EEC), and Heather classification rules
are locked in the Legacy PDF Data Import plan. Renditions
(IsRenditionFlag=Yes) are imported; same-code rows still keep latest
TargetModified.
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

from taxonomy_predicates import (
    DOCUMENT_SUBTYPE_DRAWING,
    DOCUMENT_SUBTYPE_DRAWING_DOCUMENT,
    DOCUMENT_SUBTYPE_FORM,
    DOCUMENT_SUBTYPE_PROCEDURE,
    DOCUMENT_SUBTYPE_STANDARD,
    RESERVATION_TYPE_DOCUMENT,
    RESERVATION_TYPE_DRAWING,
)

LEAF_RE = re.compile(
    r"^(?P<code>"
    r"(?P<bb>[A-Za-z0-9]{2})-(?P<aa>[A-Za-z0-9]{2})-(?P<uu>[A-Za-z0-9]{2})-"
    r"(?P<ddd>[A-Za-z0-9]{3})-(?P<sys>[A-Za-z0-9]{3})-(?P<kn>[A-Za-z0-9]{2})-"
    r"(?P<nnnn>\d{4})(?:-(?P<sss>\d{3}))?"
    r")"
    r"(?:\.(?P<office>xls|xlsx|docx|pptx|doc|ppt))?"
    r"\.pdf$",
    re.I,
)

OFFICE_EXTS = frozenset({"xls", "xlsx", "docx", "pptx", "doc", "ppt"})

IMPORT_DRAWING_DOCUMENT = "Drawing Document"
IMPORT_DRAWING = "Drawing"
IMPORT_STANDARD = "Standard"
IMPORT_PROCEDURE = "Procedure"
IMPORT_FORM = "Form"

FAMILY_DRW = "DRW"
FAMILY_STD = "STD"
FAMILY_PRC = "PRC"
FAMILY_FRM = "FRM"

STREAM_GF = "gf"
STREAM_EEC = "eec"


@dataclass
class ParsedRow:
    stream: str
    leaf: str
    url: str
    full_code: str
    coding: str
    bb: str
    aa: str
    uu: str
    ddd: str
    sys: str
    kn: str
    nnnn: int
    sss: int | None
    office: str
    excel_record_type: str
    lifecycle_status: str
    is_rendition: bool
    target_title: str
    target_modified: datetime | None
    import_type: str = ""
    reservation_type: int = 0
    document_subtype: int = 0
    sequence_family: str = ""
    parent_number: str = ""
    display_number: str = ""
    title: str = ""
    reject_category: str = ""
    reject_reason: str = ""

    @property
    def eligible(self) -> bool:
        return not self.reject_category


@dataclass
class ParseResult:
    stream: str
    source_excel: str
    total_rows: int = 0
    eligible: list[ParsedRow] = field(default_factory=list)
    rejected: list[ParsedRow] = field(default_factory=list)
    duplicates_dropped: list[dict[str, Any]] = field(default_factory=list)
    parents_needed: list[ParsedRow] = field(default_factory=list)
    sequence_seeds: dict[str, int] = field(default_factory=dict)
    missing_codes: list[dict[str, Any]] = field(default_factory=list)


def _cell(row: Mapping[str, Any], *names: str) -> Any:
    for name in names:
        if name in row and row[name] is not None:
            return row[name]
    return None


def _as_bool(val: Any) -> bool:
    if val is True or val is False:
        return bool(val)
    s = str(val or "").strip().lower()
    return s in {"1", "true", "yes", "y"}


def _parse_modified(val: Any) -> datetime | None:
    if isinstance(val, datetime):
        return val
    return None


def filename_stem_title(leaf: str) -> str:
    name = leaf
    if name.lower().endswith(".pdf"):
        name = name[:-4]
    for ext in OFFICE_EXTS:
        suffix = f".{ext}"
        if name.lower().endswith(suffix):
            name = name[: -len(suffix)]
            break
    return name


def classify_gf(sss: int | None) -> tuple[str, int, int, str]:
    if sss is None:
        return (
            IMPORT_DRAWING_DOCUMENT,
            RESERVATION_TYPE_DRAWING,
            DOCUMENT_SUBTYPE_DRAWING_DOCUMENT,
            FAMILY_DRW,
        )
    return IMPORT_DRAWING, RESERVATION_TYPE_DRAWING, DOCUMENT_SUBTYPE_DRAWING, FAMILY_DRW


def classify_eec(kn: str, sss: int | None, office: str) -> tuple[str, int, int, str]:
    # -SSS wins (including dual-ext + SSS) → Form
    if sss is not None:
        return IMPORT_FORM, RESERVATION_TYPE_DOCUMENT, DOCUMENT_SUBTYPE_FORM, FAMILY_FRM
    # ST or dual-ext without SSS → Standard
    if kn == "ST" or office:
        return IMPORT_STANDARD, RESERVATION_TYPE_DOCUMENT, DOCUMENT_SUBTYPE_STANDARD, FAMILY_STD
    return IMPORT_PROCEDURE, RESERVATION_TYPE_DOCUMENT, DOCUMENT_SUBTYPE_PROCEDURE, FAMILY_PRC


def _reject(row: ParsedRow, category: str, reason: str) -> ParsedRow:
    row.reject_category = category
    row.reject_reason = reason
    return row


def parse_excel_stream(
    stream: str,
    excel_path: Path,
    *,
    sheet_name: str = "Filtered Records",
    ref_codes: Mapping[str, set[str]] | None = None,
) -> ParseResult:
    """Parse Filtered Records into eligible / rejected sets.

    ref_codes keys: Business, Asset, Unit, Domain, System, Kind (uppercase codes).
    When None, reference validation is skipped (preflight against seed can pass sets).
    """
    import openpyxl

    result = ParseResult(stream=stream, source_excel=str(excel_path))
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not in {excel_path.name}: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}

    def g(raw: tuple[Any, ...], col: str) -> Any:
        i = idx.get(col)
        if i is None or i >= len(raw):
            return None
        return raw[i]

    by_code: dict[str, list[ParsedRow]] = defaultdict(list)

    for raw in rows_iter:
        result.total_rows += 1
        leaf = str(g(raw, "TargetFileLeafRef") or "").strip()
        url = str(g(raw, "FullTargetURL") or "").strip()
        excel_rt = str(g(raw, "RecordType") or "").strip()
        lifecycle = str(g(raw, "LifecycleStatus") or "").strip()
        title_cell = str(g(raw, "TargetTitle") or "").strip()
        rendition = _as_bool(g(raw, "IsRenditionFlag")) or _as_bool(g(raw, "IsRendition"))
        mod = _parse_modified(g(raw, "TargetModified")) or _parse_modified(g(raw, "SourceModified"))

        base = ParsedRow(
            stream=stream,
            leaf=leaf,
            url=url,
            full_code=str(g(raw, "FullCode") or "").strip(),
            coding="",
            bb="",
            aa="",
            uu="",
            ddd="",
            sys="",
            kn="",
            nnnn=0,
            sss=None,
            office="",
            excel_record_type=excel_rt,
            lifecycle_status=lifecycle,
            is_rendition=rendition,
            target_title=title_cell,
            target_modified=mod,
        )

        if not leaf.lower().endswith(".pdf"):
            result.rejected.append(_reject(base, "filename", "Not a PDF (non-.pdf extension)"))
            continue

        # Renditions (IsRenditionFlag=Yes) are imported. Same parsed code as a
        # non-rendition still dedupes to latest TargetModified below.

        if stream == STREAM_EEC:
            # Active only when LifecycleStatus column is present / populated
            if lifecycle and lifecycle.strip().lower() != "active":
                result.rejected.append(
                    _reject(
                        base,
                        "lifecycle",
                        f"LifecycleStatus={lifecycle!r} — only Active is imported",
                    )
                )
                continue

        m = LEAF_RE.match(leaf)
        if not m:
            result.rejected.append(
                _reject(
                    base,
                    "filename",
                    "Leaf is not exact Heather code[.office].pdf "
                    "(prefix/suffix text or non-Heather segments)",
                )
            )
            continue

        code = m.group("code").upper()
        office = (m.group("office") or "").lower()
        sss_raw = m.group("sss")
        sss = int(sss_raw) if sss_raw else None
        nnnn = int(m.group("nnnn"))
        bb, aa, uu = m.group("bb").upper(), m.group("aa").upper(), m.group("uu").upper()
        ddd, sys, kn = m.group("ddd").upper(), m.group("sys").upper(), m.group("kn").upper()
        coding = f"{bb}-{aa}-{uu}-{ddd}-{sys}-{kn}"
        parent_number = f"{coding}-{nnnn:04d}"
        display_number = parent_number if sss is None else f"{parent_number}-{sss:03d}"

        if stream == STREAM_GF:
            import_type, rtype, subtype, family = classify_gf(sss)
        else:
            import_type, rtype, subtype, family = classify_eec(kn, sss, office)

        title = title_cell or filename_stem_title(leaf) or display_number
        row = ParsedRow(
            stream=stream,
            leaf=leaf,
            url=url,
            full_code=(base.full_code or code),
            coding=coding,
            bb=bb,
            aa=aa,
            uu=uu,
            ddd=ddd,
            sys=sys,
            kn=kn,
            nnnn=nnnn,
            sss=sss,
            office=office,
            excel_record_type=excel_rt,
            lifecycle_status=lifecycle,
            is_rendition=rendition,
            target_title=title_cell,
            target_modified=mod,
            import_type=import_type,
            reservation_type=rtype,
            document_subtype=subtype,
            sequence_family=family,
            parent_number=parent_number,
            display_number=display_number,
            title=title,
        )
        by_code[code].append(row)

    # Latest-dedupe
    kept: list[ParsedRow] = []
    for code, items in by_code.items():
        items_sorted = sorted(
            items,
            key=lambda r: r.target_modified or datetime.min,
            reverse=True,
        )
        best = items_sorted[0]
        kept.append(best)
        for other in items_sorted[1:]:
            result.duplicates_dropped.append(
                {
                    "FullCode": other.full_code,
                    "TargetFileLeafRef": other.leaf,
                    "FullTargetURL": other.url,
                    "RecordType": other.excel_record_type,
                    "RejectCategory": "duplicate",
                    "RejectReason": (
                        f"Duplicate of {code}; kept {best.leaf} "
                        f"(newer TargetModified {best.target_modified})"
                    ),
                    "KeptLeaf": best.leaf,
                    "KeptURL": best.url,
                    "KeptModified": str(best.target_modified or ""),
                    "ThisModified": str(other.target_modified or ""),
                }
            )

    # Reference validation
    cross: dict[str, list[str]] = defaultdict(list)
    if ref_codes:
        for dim, codes in ref_codes.items():
            for c in codes:
                cross[c].append(dim)

    missing_counter: dict[tuple[str, str], int] = defaultdict(int)

    for row in kept:
        if ref_codes is not None:
            checks = [
                ("Business", row.bb),
                ("Asset", row.aa),
                ("Unit", row.uu),
                ("Domain", row.ddd),
                ("System", row.sys),
                ("Kind", row.kn),
            ]
            missing = [(d, c) for d, c in checks if c not in ref_codes.get(d, set())]
            if missing:
                reasons = []
                for dim, code in missing:
                    missing_counter[(dim, code)] += 1
                    elsewhere = [x for x in cross.get(code, []) if x != dim]
                    if elsewhere:
                        reasons.append(
                            f"{dim} code '{code}' is not in live {dim} list "
                            f"(exists as {', '.join(elsewhere)} — wrong slot)"
                        )
                    else:
                        reasons.append(
                            f"{dim} code '{code}' is not in live {dim} reference data"
                        )
                result.rejected.append(_reject(row, "reference_mismatch", "; ".join(reasons)))
                continue
        result.eligible.append(row)

    # Parents needed for sheeted types when no base PDF among eligible
    base_keys = {
        (r.parent_number, r.document_subtype)
        for r in result.eligible
        if r.sss is None
    }
    parent_seen: set[tuple[str, int]] = set()
    for r in result.eligible:
        if r.sss is None:
            continue
        key = (r.parent_number, r.document_subtype)
        if key in base_keys or key in parent_seen:
            continue
        parent_seen.add(key)
        parent = ParsedRow(
            stream=r.stream,
            leaf="",
            url="",
            full_code=r.parent_number,
            coding=r.coding,
            bb=r.bb,
            aa=r.aa,
            uu=r.uu,
            ddd=r.ddd,
            sys=r.sys,
            kn=r.kn,
            nnnn=r.nnnn,
            sss=None,
            office="",
            excel_record_type=r.excel_record_type,
            lifecycle_status="Active",
            is_rendition=False,
            target_title="",
            target_modified=None,
            import_type=r.import_type if r.import_type != IMPORT_DRAWING else IMPORT_DRAWING,
            reservation_type=r.reservation_type,
            document_subtype=r.document_subtype,
            sequence_family=r.sequence_family,
            parent_number=r.parent_number,
            display_number=r.parent_number,
            title=r.parent_number,
        )
        # Form parent stays Form; Drawing parent stays Drawing
        if r.import_type == IMPORT_FORM:
            parent.import_type = IMPORT_FORM
        elif r.import_type == IMPORT_DRAWING:
            parent.import_type = IMPORT_DRAWING
        result.parents_needed.append(parent)

    # Sequence seeds from eligible + parents (max NNNN per coding|family)
    for r in list(result.eligible) + list(result.parents_needed):
        seq_key = f"{r.coding}|{r.sequence_family}"
        if r.nnnn > result.sequence_seeds.get(seq_key, 0):
            result.sequence_seeds[seq_key] = r.nnnn

    result.missing_codes = [
        {
            "Dimension": dim,
            "MissingCode": code,
            "UniqueRecordsExcluded": n,
            "Reason": (
                f"Code '{code}' not in live {dim} reference"
                + (
                    f"; note: exists as {', '.join(x for x in cross.get(code, []) if x != dim)}"
                    if any(x != dim for x in cross.get(code, []))
                    else ""
                )
            ),
        }
        for (dim, code), n in sorted(missing_counter.items(), key=lambda t: (-t[1], t[0][0], t[0][1]))
    ]

    wb.close()
    return result


def load_seed_ref_codes(seed_dir: Path) -> dict[str, set[str]]:
    """Fallback when live Dataverse is unavailable (preflight / unit tests)."""
    mapping = {
        "Business": "business.yaml",
        "Asset": "asset.yaml",
        "Unit": "unit.yaml",
        "Domain": "domain.yaml",
        "System": "system.yaml",
        "Kind": "kind.yaml",
    }
    out: dict[str, set[str]] = {}
    for dim, fname in mapping.items():
        text = (seed_dir / fname).read_text(encoding="utf-8")
        codes: set[str] = set()
        in_rows = False
        for line in text.splitlines():
            if line.startswith("rows:"):
                in_rows = True
                continue
            if not in_rows:
                continue
            m = re.match(r"^-\s+code:\s*(.+?)\s*$", line.strip())
            if not m:
                continue
            c = m.group(1).strip()
            if len(c) >= 2 and c[0] == c[-1] and c[0] in "'\"":
                c = c[1:-1]
            codes.add(c.upper())
        out[dim] = codes
    return out


def row_to_dict(r: ParsedRow) -> dict[str, Any]:
    d = asdict(r)
    if r.target_modified:
        d["target_modified"] = r.target_modified.isoformat()
    return d
