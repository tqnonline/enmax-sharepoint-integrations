"""Write Heather-facing multi-sheet Excel import reports."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from legacy_coded_pdf_parse import ParseResult, ParsedRow


def _write_sheet(wb: Workbook, title: str, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([("" if v is None else v) for v in row])
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, len(headers[i - 1]) + 2))


def _email_draft(result: ParseResult, env: str, mode: str) -> list[list[Any]]:
    eligible = len(result.eligible)
    rejected = len(result.rejected)
    dups = len(result.duplicates_dropped)
    by_type: dict[str, int] = {}
    for r in result.eligible:
        by_type[r.import_type] = by_type.get(r.import_type, 0) + 1
    by_cat: dict[str, int] = {}
    for r in result.rejected:
        by_cat[r.reject_category] = by_cat.get(r.reject_category, 0) + 1

    lines = [
        ["Field", "Value"],
        ["To", "Heather"],
        ["Subject", f"Legacy PDF import report — {result.stream.upper()} — {env} ({mode})"],
        ["Environment", env],
        ["Mode", mode],
        ["Stream", result.stream.upper()],
        ["Source", result.source_excel],
        ["Generated (UTC)", datetime.now(timezone.utc).isoformat()],
        [],
        ["Body", ""],
        ["", f"We processed the {result.stream.upper()} coded-PDF migration extract."],
        ["", f"Eligible for import / imported: {eligible}"],
        ["", f"Not imported (policy exclusions): {rejected}"],
        ["", f"Duplicate leaves dropped (kept latest): {dups}"],
        ["", "Imported by type:"],
    ]
    for k, v in sorted(by_type.items()):
        lines.append(["", f"  - {k}: {v}"])
    lines.append(["", "Not imported by reason category:"])
    for k, v in sorted(by_cat.items(), key=lambda t: -t[1]):
        lines.append(["", f"  - {k}: {v}"])
    lines.append(["", "See workbook sheets for full row detail and RejectReason text."])
    lines.append(["", "Next: Search spot-check; proceed UAT when checks pass."])
    return lines


def write_heather_workbook(
    result: ParseResult,
    out_path: Path,
    *,
    env: str = "dev",
    mode: str = "preflight",
    applied_ids: dict[str, str] | None = None,
) -> Path:
    """Write multi-sheet Heather report. applied_ids maps display_number|subtype -> guid."""
    applied_ids = applied_ids or {}
    wb = Workbook()
    # Remove default sheet after creating content
    default = wb.active
    wb.remove(default)

    # Read Me
    readme = [
        ["Legacy coded PDF import report"],
        ["Stream", result.stream.upper()],
        ["Environment", env],
        ["Mode", mode],
        ["Source Excel", result.source_excel],
        ["Generated (UTC)", datetime.now(timezone.utc).isoformat()],
        [],
        ["Policy"],
        ["Exact Heather leaf only (no prefix/suffix text)"],
        ["PDF only; dual-ext .xls.pdf/.docx.pdf etc. allowed"],
        ["Duplicates: keep latest TargetModified"],
        ["Renditions included (IsRenditionFlag=Yes)"],
        ["EEC: Active LifecycleStatus only"],
        ["Unknown Business/Asset/Unit/Domain/System/Kind excluded"],
        ["EEC: -SSS => Form; ST or dual-ext (no SSS) => Standard; else Procedure"],
        ["GF: no SSS => Drawing Document; with SSS => Drawing sheet"],
        ["Parents created when only sheet PDFs exist"],
    ]
    ws = wb.create_sheet("Read Me", 0)
    for row in readme:
        ws.append(row)

    # Summary
    by_type: dict[str, int] = {}
    for r in result.eligible:
        by_type[r.import_type] = by_type.get(r.import_type, 0) + 1
    by_cat: dict[str, int] = {}
    for r in result.rejected:
        by_cat[r.reject_category] = by_cat.get(r.reject_category, 0) + 1
    summary_rows = [
        ["Metric", "Count"],
        ["Filtered Records rows", result.total_rows],
        ["Eligible / imported rows", len(result.eligible)],
        ["Parents without base PDF", len(result.parents_needed)],
        ["Not imported", len(result.rejected)],
        ["Duplicates dropped", len(result.duplicates_dropped)],
        ["Sequence keys seeded", len(result.sequence_seeds)],
        [],
        ["Eligible by ImportType", "Count"],
        *[[k, v] for k, v in sorted(by_type.items())],
        [],
        ["Not imported by category", "Count"],
        *[[k, v] for k, v in sorted(by_cat.items(), key=lambda t: -t[1])],
    ]
    _write_sheet(wb, "Summary", summary_rows[0], summary_rows[1:])

    # Email Draft
    draft = _email_draft(result, env, mode)
    _write_sheet(wb, "Email Draft", draft[0], draft[1:])

    def imported_rows(rows: list[ParsedRow]) -> list[list[Any]]:
        out = []
        for r in rows:
            key = f"{r.display_number}|{r.document_subtype}"
            out.append(
                [
                    r.full_code,
                    r.display_number,
                    r.parent_number,
                    r.import_type,
                    r.excel_record_type,
                    r.coding,
                    r.nnnn,
                    r.sss if r.sss is not None else "",
                    r.bb,
                    r.aa,
                    r.uu,
                    r.ddd,
                    r.sys,
                    r.kn,
                    r.office,
                    r.url,
                    r.title,
                    r.reservation_type,
                    r.document_subtype,
                    r.sequence_family,
                    applied_ids.get(key, ""),
                    env,
                ]
            )
        return out

    headers_imp = [
        "FullCode",
        "DisplayNumber",
        "ParentNumber",
        "ImportType",
        "ExcelRecordType",
        "CodingSequence",
        "NNNN",
        "SSS",
        "BB",
        "AA",
        "UU",
        "DDD",
        "SYS",
        "KN",
        "OfficeStem",
        "FullTargetURL",
        "Title",
        "ReservationType",
        "DocumentSubtype",
        "SequenceFamily",
        "DataverseId",
        "Environment",
    ]

    docs = [r for r in result.eligible if r.import_type in {"Drawing Document", "Standard", "Procedure"}]
    sheets = [r for r in result.eligible if r.sss is not None]
    _write_sheet(wb, "Imported Bases", headers_imp, imported_rows(docs))
    _write_sheet(wb, "Imported Sheets", headers_imp, imported_rows(sheets))
    _write_sheet(wb, "Imported Parents", headers_imp, imported_rows(result.parents_needed))

    _write_sheet(
        wb,
        "Number Sequences Seeded",
        ["SequenceKey", "LastIssued", "SeedValue", "Reason"],
        [
            [k, v, 0, "max NNNN among imported rows for this coding|family"]
            for k, v in sorted(result.sequence_seeds.items())
        ],
    )

    not_headers = [
        "FullCode",
        "TargetFileLeafRef",
        "FullTargetURL",
        "ExcelRecordType",
        "LifecycleStatus",
        "RejectCategory",
        "RejectReason",
        "BB",
        "AA",
        "UU",
        "DDD",
        "SYS",
        "KN",
    ]

    def rej_rows(pred) -> list[list[Any]]:
        return [
            [
                r.full_code,
                r.leaf,
                r.url,
                r.excel_record_type,
                r.lifecycle_status,
                r.reject_category,
                r.reject_reason,
                r.bb,
                r.aa,
                r.uu,
                r.ddd,
                r.sys,
                r.kn,
            ]
            for r in result.rejected
            if pred(r)
        ]

    _write_sheet(wb, "Not Imported - All", not_headers, rej_rows(lambda r: True))
    _write_sheet(
        wb,
        "Not Imported - Filename",
        not_headers,
        rej_rows(lambda r: r.reject_category == "filename"),
    )
    _write_sheet(
        wb,
        "Not Imported - Lifecycle",
        not_headers,
        rej_rows(lambda r: r.reject_category == "lifecycle"),
    )
    _write_sheet(
        wb,
        "Not Imported - Rendition",
        not_headers,
        rej_rows(lambda r: r.reject_category == "rendition"),
    )
    _write_sheet(
        wb,
        "Not Imported - Reference",
        not_headers,
        rej_rows(lambda r: r.reject_category == "reference_mismatch"),
    )

    dup_headers = [
        "FullCode",
        "TargetFileLeafRef",
        "FullTargetURL",
        "RecordType",
        "RejectCategory",
        "RejectReason",
        "KeptLeaf",
        "KeptURL",
        "KeptModified",
        "ThisModified",
    ]
    _write_sheet(
        wb,
        "Not Imported - Duplicates",
        dup_headers,
        [[d.get(h, "") for h in dup_headers] for d in result.duplicates_dropped],
    )

    _write_sheet(
        wb,
        "Missing Codes Rollup",
        ["Dimension", "MissingCode", "UniqueRecordsExcluded", "Reason"],
        [
            [m["Dimension"], m["MissingCode"], m["UniqueRecordsExcluded"], m["Reason"]]
            for m in result.missing_codes
        ],
    )

    if result.stream == "eec":
        class_counts: dict[tuple[str, str], int] = {}
        for r in result.eligible:
            key = (r.excel_record_type or "(blank)", r.import_type)
            class_counts[key] = class_counts.get(key, 0) + 1
        _write_sheet(
            wb,
            "Classification",
            ["ExcelRecordType", "ImportType", "Count"],
            [[a, b, n] for (a, b), n in sorted(class_counts.items(), key=lambda t: -t[1])],
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
